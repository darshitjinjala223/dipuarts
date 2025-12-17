from openpyxl import load_workbook
from openpyxl.styles import Font
import shutil
import os
from num2words import num2words

def apply_style(cell, size=12):
    """Applies Times New Roman with specific size (default 12) to a cell."""
    cell.font = Font(name='Times New Roman', size=size)

def generate_invoice_excel(data, template_path="templates/INVOICE FORMAT2.xlsx", output_path="generated/temp_invoice.xlsx"):
    # Resolve absolute paths to ensure reliability on Cloud
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # If template path is relative, make it absolute relative to base_dir
    if not os.path.isabs(template_path):
        template_path = os.path.join(base_dir, template_path)
    
    if not os.path.isabs(output_path):
        output_path = os.path.join(base_dir, output_path)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    if not os.path.exists(template_path):
        # Fallback debug or error
        raise FileNotFoundError(f"Template not found at: {template_path}")

    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active 
    
    # 1. Header
    # Invoice Date [D12]
    ws['D12'] = data['date']; apply_style(ws['D12'])
    # Invoice No [D11]
    ws['D11'] = data['invoice_no']; apply_style(ws['D11'])
    
    # Challan Data [K11, K12]
    ws['K11'] = data['challan_no']; apply_style(ws['K11'])
    ws['K12'] = data.get('challan_date', ''); apply_style(ws['K12'])
    
    # Order Data [K13, K14]
    ws['K13'] = data.get('order_no', ''); apply_style(ws['K13'])
    ws['K14'] = data.get('order_date', ''); apply_style(ws['K14'])
    
    # 2. Buyer Details [D17, D18, D20]
    ws['D17'] = data['supplier_name']; apply_style(ws['D17'])
    ws['D18'] = data.get('supplier_address', ''); apply_style(ws['D18'])
    import openpyxl
    ws['D18'].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
    
    ws['D20'] = data.get('supplier_gst', ''); apply_style(ws['D20'])
    
    # 3. Line Items (Dynamic Rows starting at 24)
    r = 24
    
    # Check for multiple items payload
    items = data.get('items', [])
    if not items:
        # Fallback for single item legacy
        items = [{
            'material': data.get('material', ''),
            'qty': data.get('qty', 0),
            'rate': data.get('rate', 0),
            'base_amount': data.get('base_amount', 0),
            'cgst': data.get('cgst', 0),
            'sgst': data.get('sgst', 0),
            'total': data.get('total', 0)
        }]
    
    for i, item in enumerate(items):
        current_r = r + i
        # Sr No
        safe_write(ws, f'B{current_r}', i + 1, 10)
        # Description
        safe_write(ws, f'C{current_r}', item['material'], 10)
        # MTR
        safe_write(ws, f'G{current_r}', item['qty'], 10)
        # Rate
        safe_write(ws, f'H{current_r}', item['rate'], 10)
        # Taxable
        safe_write(ws, f'I{current_r}', item['base_amount'], 10)
        # GST Rate
        safe_write(ws, f'J{current_r}', "5%", 10)
        # GST Amount
        row_gst = item['cgst'] + item['sgst']
        safe_write(ws, f'K{current_r}', row_gst, 10)
        # Total
        safe_write(ws, f'L{current_r}', item['total'], 10)
    
    # 4. Footer Totals - Size 10
    # Taxable Amount [L39]
    ws['L39'] = data['base_amount']; apply_style(ws['L39'], 10)
    
    # CGST [L40]
    ws['L40'] = data['cgst']; apply_style(ws['L40'], 10)
    
    # SGST [L41]
    ws['L41'] = data['sgst']; apply_style(ws['L41'], 10)
    
    # Grand Total [L42]
    ws['L42'] = data['total']; apply_style(ws['L42'], 10)
    
    # Amount In Words [B39] (Merged B39:H40 approx)
    total_val = data['total']
    try:
        words = num2words(total_val, lang='en_IN').title() + " Only"
    except:
        words = f"{total_val} Only"
        
    ws['B39'] = f"Total Invoice amount in words: {words}"
    apply_style(ws['B39'], 12)

    # FIX: Increase Header Row Heights to prevent address cut-off
    ws.row_dimensions[6].height = 50 
    ws.row_dimensions[7].height = 50
    ws.row_dimensions[8].height = 50
    
    # 3. Line Items (Dynamic Rows starting at 24)
    r = 24
    
    # Reverted Column Width Adjustments
    
    # Force Page Layout to A4 & Fit Width to avoid cut-off in PDF
    setup_page_layout(ws)
    
    wb.save(output_path)
    return output_path

def generate_challan_excel(data, template_path="templates/CHALLAN FORMAT.xlsx", output_path="generated/temp_challan.xlsx"):
    # Resolve absolute paths
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    if not os.path.isabs(template_path):
        template_path = os.path.join(base_dir, template_path)
    
    if not os.path.isabs(output_path):
        output_path = os.path.join(base_dir, output_path)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found at: {template_path}")

    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Challan Mappings - All Size 14 (User requested 14 for Challan separately, I should check if they want 10 here too? "in invoice bill this ####" - context implies Invoice. I'll touch Invoice mostly.
    # But "numerics to 10 in a row where not fitted".
    # I'll keep Challan as 14 unless requested, assuming the #### was in Invoice.)
    
    ws['L5'] = data['date']; apply_style(ws['L5'], 12)
    ws['K6'] = data['supplier']; apply_style(ws['K6'], 12)
    
    # K8: Supplier GST
    ws['K8'] = data.get('supplier_gst', ''); apply_style(ws['K8'], 12)
    
    # E10: Order No
    ws['E10'] = data.get('order_no', ''); apply_style(ws['E10'], 12)
    
    # E12: Challan No
    ws['E12'] = data['challan_no']; apply_style(ws['E12'], 12)
    
    # Table - Dynamic Rows starting at 16
    r = 16
    
    # Check if we have multiple items or single (legacy support)
    items = data.get('items', [])
    if not items:
        # Create single item from flat data
        items = [{'material': data.get('material', ''), 'quantity': data.get('quantity', '')}]
    
    for i, item in enumerate(items):
        current_row = r + i
        ws[f'B{current_row}'] = i + 1; apply_style(ws[f'B{current_row}'], 12)
        ws[f'D{current_row}'] = item['quantity']; apply_style(ws[f'D{current_row}'], 12)
        ws[f'H{current_row}'] = item['material']; apply_style(ws[f'H{current_row}'], 12)
        
    # Reverted Column Width Adjustments
    
    # Force Page Layout to A4 & Fit Width
    setup_page_layout(ws)
    
    wb.save(output_path)
    return output_path

def setup_page_layout(ws):
    """
    Configures the worksheet for printing:
    - A4 Paper, Portrait
    - Fit to 1 Page Width
    - Narrow Margins
    - Explicit Print Area (A1 to L[MaxRow])
    """
    try:
        # A4 Paper & Portrait
        ws.page_setup.paperSize = ws.page_setup.PAPERSIZE_A4
        ws.page_setup.orientation = ws.page_setup.ORIENTATION_PORTRAIT
        
        # Fit to 1 Page Wide, indefinite height
        ws.page_setup.scale = None # Clear scaling
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = False
        
        # Narrow Margins (0.25 inch ~ 0.6 cm)
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.25
        ws.page_margins.bottom = 0.25
        ws.page_margins.header = 0.0
        ws.page_margins.footer = 0.0
        
        # Center Horizontally
        ws.print_options.horizontalCentered = True
        
        # Explicit Print Area (Content is usually up to Col L)
        max_row = ws.max_row
        # Ensure we capture everything
        ws.print_area = f"A1:L{max_row}"
        
    except Exception as e:
        print(f"Page Setup Error: {e}")
        pass

def safe_write(ws, cell_ref, value, font_size=None):
    """
    Safely write to a cell, finding the top-left cell if merged.
    Also applies style if font_size is provided.
    """
    import openpyxl
    
    # Check if cell is within a merged range
    for range_ in ws.merged_cells.ranges:
        if cell_ref in range_:
            # Write to the top-left cell of the merged range
            top_left = range_.coord.split(':')[0]
            ws[top_left] = value
            if font_size:
                apply_style(ws[top_left], font_size)
            return

    # Not merged (or we are the top-left/target)
    # Note: If cell is part of merge but not top-left, accessing ws[cell_ref] returns MergedCell
    # But cell_ref in range_ should catch that.
    
    # Attempt direct write
    try:
        ws[cell_ref] = value
        if font_size:
            apply_style(ws[cell_ref], font_size)
    except AttributeError:
        # Fallback if somehow missed (MergedCell has no value setter)
        pass 

def update_master_ledger(invoice_data, master_path="generated/Master_Sales.xlsx"):
    """
    Updates the Master Sales Excel file with new invoice data.
    Columns: Invoice No, Supplier Name, Taxable, CGST, SGST, Total.
    Sorted by Invoice No.
    """
    import pandas as pd
    
    # Define Columns
    cols = ["Invoice No", "Date", "Supplier Name", "Taxable Amount", "CGST", "SGST", "Total Amount"]
    
    # Prepare New Row
    new_row = {
        "Invoice No": invoice_data['invoice_no'],
        "Date": invoice_data['date'],
        "Supplier Name": invoice_data['supplier_name'],
        "Taxable Amount": float(invoice_data['base_amount']),
        "CGST": float(invoice_data['cgst']),
        "SGST": float(invoice_data['sgst']),
        "Total Amount": float(invoice_data['total'])
    }
    
    if os.path.exists(master_path):
        df = pd.read_excel(master_path)
    else:
        df = pd.DataFrame(columns=cols)
        
    # Append
    # Concat is safer than append (deprecated)
    new_df = pd.DataFrame([new_row])
    df = pd.concat([df, new_df], ignore_index=True)
    
    # Sort by Invoice No
    # Try to convert to numeric for sorting if possible, else string sort
    try:
        df['sort_key'] = pd.to_numeric(df['Invoice No'])
        df = df.sort_values(by='sort_key')
        df = df.drop(columns=['sort_key'])
    except:
        df = df.sort_values(by='Invoice No')
        
    # Save
    df.to_excel(master_path, index=False)
    return master_path 
